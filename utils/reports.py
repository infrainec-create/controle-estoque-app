import io
import pandas as pd
import numpy as np
import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from database.connection import get_conn

def formatar_aba_excel(ws, title_color="1E3A8A"):
    """
    Aplica formatação visual premium (cabeçalhos coloridos, auto-ajuste e bordas) a uma aba do Excel.
    """
    # Cabeçalho premium (Fonte branca, negrito, centralizado, preenchimento azul escuro)
    header_fill = PatternFill(start_color=title_color, end_color=title_color, fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    # Formatação dos cabeçalhos (Linha 1)
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border
    
    # Formatação das linhas de dados
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            # Se for texto, alinha à esquerda, se for número/data, centraliza
            if isinstance(cell.value, str):
                cell.alignment = align_left
            else:
                cell.alignment = align_center
                
    # Auto-ajuste de largura de colunas
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_to_check = cell.value or ''
            # Se for float formatado em dinheiro, adiciona tamanho para R$
            val_str = str(val_to_check)
            if isinstance(val_to_check, float) and cell.number_format and "R$" in cell.number_format:
                val_str = f"R$ {val_to_check:,.2f}"
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

def gerar_excel_estoque(df):
    """
    Gera arquivo Excel (.xlsx) premium contendo a posição de estoque, valuation de ativos e métricas de suprimentos.
    """
    from utils.consumption import processar_consumo_produtos
    
    # 1. Carrega método da sessão ou usa padrão
    metodo = "movimentacoes"
    janela_dias = 30
    try:
        import streamlit as st
        if "metodo_consumo" in st.session_state:
            metodo = st.session_state["metodo_consumo"]
    except Exception:
        pass
        
    # Carrega fatores de segurança por setor configurados no banco
    fatores_setor = {}
    padroes = {"Limpeza": 1.1, "Copa": 1.1, "EPI": 1.2, "Escritório": 1.1, "Geral": 1.1}
    try:
        with get_conn() as conn:
            rows_f = conn.execute("SELECT chave, valor FROM configuracoes WHERE chave LIKE 'fator_seguranca_%'").fetchall()
            for k, v in rows_f:
                setor_nome = k.replace("fator_seguranca_", "")
                fatores_setor[setor_nome] = float(v)
    except Exception:
        pass

    buffer = io.BytesIO()
    
    # Prepara dados limpos para exportação
    df_export = df.copy()
    if "criticidade" not in df_export.columns:
        df_export["criticidade"] = "Y"
        
    df_export["valor_total"] = df_export["saldo_atual"] * df_export["valor_unitario"]
    
    # Curva ABC
    total_valuation = df_export["valor_total"].sum()
    df_abc = df_export.sort_values(by="valor_total", ascending=False).copy()
    classes_map = {}
    if total_valuation > 0:
        df_abc["valor_acumulado"] = df_abc["valor_total"].cumsum()
        df_abc["perc_acumulado"] = (df_abc["valor_acumulado"] / total_valuation) * 100
        for _, row in df_abc.iterrows():
            perc = row["perc_acumulado"]
            classes_map[row["id"]] = "A" if perc <= 80 else ("B" if perc <= 95 else "C")
    else:
        classes_map = {row["id"]: "C" for _, row in df_export.iterrows()}
        
    df_export["Classe_ABC"] = df_export["id"].map(classes_map).fillna("C")
    
    # Calcular Consumo Diário e histórico de 3 semanas via função oficial
    df_export = processar_consumo_produtos(df_export, metodo, janela_dias)
    
    # Ponto de Pedido / Ressuprimento alinhado ao dashboard
    def obter_fator_setor(row):
        cat = row["categoria"]
        return fatores_setor.get(cat, padroes.get(cat, 1.1))
        
    df_export["Fator_Seguranca"] = df_export.apply(obter_fator_setor, axis=1)
    df_export["Estoque_Seguranca"] = np.maximum(df_export["estoque_minimo"], np.ceil(df_export["consumo_diario"] * df_export["lead_time"] * df_export["Fator_Seguranca"]).astype(int))
    df_export["Consumo_Lead_Time"] = np.ceil(df_export["consumo_diario"] * df_export["lead_time"]).astype(int)
    df_export["Ponto_Pedido"] = df_export["Consumo_Lead_Time"] + df_export["Estoque_Seguranca"]
    
    # Runway (Cobertura em dias)
    df_export["Runway"] = 999
    mask_consumo = df_export["consumo_diario"] > 0
    df_export.loc[mask_consumo, "Runway"] = (df_export.loc[mask_consumo, "saldo_atual"] / df_export.loc[mask_consumo, "consumo_diario"]).astype(int)
    df_export["Runway_Txt"] = df_export["Runway"].apply(lambda x: "Sem consumo" if x == 999 else f"{x} dias")
    
    # Sugestão de Compra
    df_export["Minimo_Ideal"] = df_export["Estoque_Seguranca"]
    df_export["Sugestao_Compra"] = (df_export["Minimo_Ideal"] - df_export["saldo_atual"]).clip(lower=0)
    df_export["Custo_Compra"] = df_export["Sugestao_Compra"] * df_export["valor_unitario"]
    
    # Status
    def calc_status(row):
        if row["saldo_atual"] <= 0:
            return "🔴 Ruptura"
        if row["saldo_atual"] < row["estoque_minimo"]:
            return "🔴 Crítico"
        if row["saldo_atual"] <= row["Ponto_Pedido"]:
            return "🟠 Ponto de Pedido"
        return "🟢 OK"
    df_export["Status"] = df_export.apply(calc_status, axis=1)

    # Renomeia colunas para cabeçalhos amigáveis em português
    df_export = df_export.rename(columns={
        "id": "ID Produto",
        "nome": "Insumo / Item",
        "saldo_atual": "Saldo Atual",
        "estoque_minimo": "Estoque Mínimo (Membro)",
        "valor_unitario": "Valor Unitário (R$)",
        "categoria": "Setor / Categoria",
        "lead_time": "Lead Time (Dias)",
        "consumo_diario": "Consumo Diário (Médio)",
        "Runway_Txt": "Cobertura (Runway)",
        "Classe_ABC": "Classe Financeira (ABC)",
        "criticidade": "Criticidade Operacional (XYZ)",
        "Ponto_Pedido": "Ponto de Ressuprimento",
        "Minimo_Ideal": "Estoque de Segurança",
        "Sugestao_Compra": "Sugestão de Reposição (Qtd)",
        "Custo_Compra": "Custo de Reposição (R$)",
        "valor_total": "Valor Total Ativo (R$)"
    })
    
    # Ordena colunas para ficar muito organizado
    colunas_ordenadas = [
        "ID Produto", "Setor / Categoria", "Insumo / Item", "Status",
        "Saldo Atual", "Estoque Mínimo (Membro)", "Estoque de Segurança", "Ponto de Ressuprimento", "Lead Time (Dias)",
        "Consumo Diário (Médio)", "Cobertura (Runway)", "Classe Financeira (ABC)", "Criticidade Operacional (XYZ)",
        "Valor Unitário (R$)", "Valor Total Ativo (R$)", "Sugestão de Reposição (Qtd)", "Custo de Reposição (R$)"
    ]
    df_export = df_export[colunas_ordenadas]
    
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df_export.to_excel(writer, sheet_name='Valuation de Estoque', index=False)
        
        worksheet = writer.sheets['Valuation de Estoque']
        
        # Formata colunas de valor
        for row in range(2, worksheet.max_row + 1):
            cell_unit = worksheet.cell(row=row, column=14) 
            cell_total = worksheet.cell(row=row, column=15) 
            cell_custo = worksheet.cell(row=row, column=17) 
            cell_unit.number_format = 'R$ #,##0.00'
            cell_total.number_format = 'R$ #,##0.00'
            cell_custo.number_format = 'R$ #,##0.00'
            
        # Adiciona linha de totais na base
        max_row = worksheet.max_row
        totals_row = max_row + 2
        
        # Estilo de totalização
        font_total = Font(name="Calibri", size=11, bold=True, color="000000")
        fill_total = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        border_total = Border(
            top=Side(style='thin', color='9CA3AF'),
            bottom=Side(style='double', color='000000')
        )
        
        worksheet.cell(row=totals_row, column=3, value="TOTAIS DO WMS:").font = font_total
        worksheet.cell(row=totals_row, column=3).alignment = Alignment(horizontal="right")
        
        # Fórmulas de totalização do Excel
        cell_sum_saldo = worksheet.cell(row=totals_row, column=5, value=f"=SUM(E2:E{max_row})")
        cell_sum_val = worksheet.cell(row=totals_row, column=15, value=f"=SUM(O2:O{max_row})")
        cell_sum_sug = worksheet.cell(row=totals_row, column=16, value=f"=SUM(P2:P{max_row})")
        cell_sum_custo = worksheet.cell(row=totals_row, column=17, value=f"=SUM(Q2:Q{max_row})")
        
        for c_idx in range(1, 18):
            c_cell = worksheet.cell(row=totals_row, column=c_idx)
            c_cell.font = font_total
            c_cell.fill = fill_total
            c_cell.border = border_total
            
        cell_sum_saldo.alignment = Alignment(horizontal="center")
        cell_sum_val.number_format = 'R$ #,##0.00'
        cell_sum_val.alignment = Alignment(horizontal="center")
        cell_sum_sug.alignment = Alignment(horizontal="center")
        cell_sum_custo.number_format = 'R$ #,##0.00'
        cell_sum_custo.alignment = Alignment(horizontal="center")
        
        formatar_aba_excel(worksheet, title_color="1E3A8A")
        
    buffer.seek(0)
    return buffer.getvalue()

def gerar_excel_movimentacoes(mv):
    """
    Gera o extrato de movimentações do sistema em formato Excel formatado.
    """
    buffer = io.BytesIO()
    
    df_export = mv.copy()
    df_export = df_export.rename(columns={
        "id": "ID Lançamento",
        "produto": "Insumo / Item",
        "data_hora": "Data/Hora",
        "tipo": "Tipo Operação",
        "quantidade": "Qtd. Movimentada",
        "saldo_resultante": "Saldo Resultante",
        "observacao": "Detalhes / Motivo"
    })
    
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df_export.to_excel(writer, sheet_name='Histórico de Fluxo', index=False)
        formatar_aba_excel(writer.sheets['Histórico de Fluxo'], title_color="0F766E") # Cor verde azulado
        
    buffer.seek(0)
    return buffer.getvalue()

def gerar_excel_auditoria(logs):
    """
    Gera o log completo de auditoria do sistema em formato Excel formatado para compliance.
    """
    buffer = io.BytesIO()
    
    df_export = logs.copy()
    df_export = df_export.rename(columns={
        "id": "ID Registro",
        "usuario": "Operador",
        "acao": "Tipo de Ação",
        "data_hora": "Data/Hora",
        "detalhes": "Detalhes Operacionais"
    })
    
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df_export.to_excel(writer, sheet_name='Logs de Auditoria', index=False)
        formatar_aba_excel(writer.sheets['Logs de Auditoria'], title_color="374151") # Cor cinza escuro
        
    buffer.seek(0)
    return buffer.getvalue()

def gerar_html_pdf_estoque(df, mv, logs, metodo=None, janela_dias=30):
    """
    Compila um relatório executivo de alta fidelidade visual (HTML) otimizado para salvamento em PDF / Impressão.
    """
    from utils.consumption import processar_consumo_produtos
    
    # 1. Carregar método da sessão se não fornecido
    if metodo is None:
        try:
            import streamlit as st
            metodo = st.session_state.get("metodo_consumo", "movimentacoes")
        except Exception:
            metodo = "movimentacoes"
            
    # Carregar fatores de segurança por setor configurados no banco
    fatores_setor = {}
    padroes = {"Limpeza": 1.1, "Copa": 1.1, "EPI": 1.2, "Escritório": 1.1, "Geral": 1.1}
    try:
        with get_conn() as conn:
            rows_f = conn.execute("SELECT chave, valor FROM configuracoes WHERE chave LIKE 'fator_seguranca_%'").fetchall()
            for k, v in rows_f:
                setor_nome = k.replace("fator_seguranca_", "")
                fatores_setor[setor_nome] = float(v)
    except Exception:
        pass

    # 2. Calcular Valuation, Cobertura, Classe ABC, Ponto de Pedido e Sugestões
    df_calc = df.copy()
    if "criticidade" not in df_calc.columns:
        df_calc["criticidade"] = "Y"
        
    df_calc["valor_total"] = df_calc["saldo_atual"] * df_calc["valor_unitario"]
    
    # Curva ABC (Classe A, B, C)
    total_valuation = df_calc["valor_total"].sum()
    df_abc = df_calc.sort_values(by="valor_total", ascending=False).copy()
    classes_map = {}
    if total_valuation > 0:
        df_abc["valor_acumulado"] = df_abc["valor_total"].cumsum()
        df_abc["perc_acumulado"] = (df_abc["valor_acumulado"] / total_valuation) * 100
        for _, row in df_abc.iterrows():
            perc = row["perc_acumulado"]
            classes_map[row["id"]] = "A" if perc <= 80 else ("B" if perc <= 95 else "C")
    else:
        classes_map = {row["id"]: "C" for _, row in df_calc.iterrows()}
        
    df_calc["Classe_ABC"] = df_calc["id"].map(classes_map).fillna("C")
    
    # Processar consumo via função oficial do sistema (alinhada ao painel)
    df_calc = processar_consumo_produtos(df_calc, metodo, janela_dias)
    
    # Ponto de Pedido / Ressuprimento alinhado ao dashboard
    def obter_fator_setor(row):
        cat = row["categoria"]
        return fatores_setor.get(cat, padroes.get(cat, 1.1))
        
    df_calc["Fator_Seguranca"] = df_calc.apply(obter_fator_setor, axis=1)
    df_calc["Estoque_Seguranca"] = np.maximum(df_calc["estoque_minimo"], np.ceil(df_calc["consumo_diario"] * df_calc["lead_time"] * df_calc["Fator_Seguranca"]).astype(int))
    df_calc["Consumo_Lead_Time"] = np.ceil(df_calc["consumo_diario"] * df_calc["lead_time"]).astype(int)
    df_calc["Ponto_Pedido"] = df_calc["Consumo_Lead_Time"] + df_calc["Estoque_Seguranca"]
    
    # Runway (Cobertura em dias)
    df_calc["Runway"] = 999
    mask_consumo = df_calc["consumo_diario"] > 0
    df_calc.loc[mask_consumo, "Runway"] = (df_calc.loc[mask_consumo, "saldo_atual"] / df_calc.loc[mask_consumo, "consumo_diario"]).astype(int)
    
    # Minimo Ideal e Sugestão de Compra
    df_calc["Minimo_Ideal"] = df_calc["Estoque_Seguranca"]
    df_calc["Sugestao_Compra"] = (df_calc["Minimo_Ideal"] - df_calc["saldo_atual"]).clip(lower=0)
    df_calc["Custo_Compra"] = df_calc["Sugestao_Compra"] * df_calc["valor_unitario"]
    
    # Status
    def calc_status(row):
        if row["saldo_atual"] <= 0:
            return "Ruptura"
        if row["saldo_atual"] < row["estoque_minimo"]:
            return "Crítico"
        if row["saldo_atual"] <= row["Ponto_Pedido"]:
            return "Ponto de Pedido"
        return "OK"
        
    df_calc["Status"] = df_calc.apply(calc_status, axis=1)
    
    # 3. Métricas de Inventário (Acuracidade e Perdas)
    total_contagens = 0
    contagens_corretas = 0
    perdas_financeiras = 0.0
    
    if not mv.empty:
        contagens = mv[mv["tipo"] == "Contagem"]
        total_contagens = len(contagens)
        if total_contagens > 0:
            contagens_corretas = len(contagens[contagens["quantidade"] == 0])
            
            # Perdas financeiras: ajustes negativos * preço
            precos_dict = dict(zip(df_calc["nome"], df_calc["valor_unitario"]))
            for _, r in contagens.iterrows():
                diff = r["quantidade"]
                if diff < 0:
                    prod = r["produto"]
                    p_unit = precos_dict.get(prod, 0.0)
                    perdas_financeiras += abs(diff) * p_unit
                    
    ira = (contagens_corretas / total_contagens * 100) if total_contagens > 0 else 100.0

    # 4. Resumo Executivo para Ação Imediata
    urgente_comprar = []
    monitorar_aprovisionar = []
    investigar = []
    
    for _, row in df_calc.iterrows():
        status = row["Status"]
        nome = row["nome"]
        saldo = row["saldo_atual"]
        runway = row["Runway"]
        minimo = row["estoque_minimo"]
        pmp = row["valor_unitario"]
        
        if status in ["Ruptura", "Crítico"] and row["Sugestao_Compra"] > 0:
            urgente_comprar.append(f"<b>{nome}</b> (Saldo: {saldo} un, Compra Recomendada: {int(row['Sugestao_Compra'])} un)")
        elif status == "Ponto de Pedido" or saldo == 1:
            monitorar_aprovisionar.append(f"<b>{nome}</b> (Saldo: {saldo} un, Cobertura: {runway if runway != 999 else 'N/A'} dias)")
        
        # Investigar excesso
        if runway != 999 and runway > 90 and (saldo * pmp) > 100.0:
            investigar.append(f"<b>{nome}</b>: Cobertura de {runway} dias. Saldo: {saldo} un (Mín: {minimo} un). Capital imobilizado: R$ {saldo * pmp:,.2f}.")
        # Investigar anomalias de preço
        if pmp > 40.0 and any(kw in nome for kw in ["Papel", "Pano", "Detergente", "Esponja", "Café"]):
            investigar.append(f"<b>{nome}</b>: Verificar preço unitário cadastrado (R$ {pmp:,.2f}/un).")

    # 5. Ordenação por Criticidade na Posição de Estoque
    status_order = {"Ruptura": 0, "Crítico": 1, "Ponto de Pedido": 2, "OK": 3}
    df_calc["status_priority"] = df_calc["Status"].map(status_order)
    df_estoque_ordenado = df_calc.sort_values(by=["status_priority", "categoria", "nome"]).copy()

    # 6. Gráficos em SVG (Valuation por Categoria)
    cat_val = df_calc.groupby("categoria")["valor_total"].sum().reset_index()
    cat_val = cat_val[cat_val["valor_total"] > 0].sort_values(by="valor_total", ascending=False)
    
    svg_chart = ""
    if not cat_val.empty:
        svg_width = 500
        svg_height = len(cat_val) * 35 + 40
        chart_rows = ""
        max_val = cat_val["valor_total"].max()
        for idx, (_, r) in enumerate(cat_val.iterrows()):
            cat = r["categoria"]
            val = r["valor_total"]
            bar_width = int((val / max_val) * 250) if max_val > 0 else 0
            y_pos = idx * 35 + 25
            chart_rows += f"""
            <text x="10" y="{y_pos + 14}" font-family="Inter" font-size="11px" font-weight="500" fill="#374151">{cat}</text>
            <rect x="150" y="{y_pos}" width="{bar_width}" height="18" rx="3" fill="#1E3A8A" />
            <text x="{160 + bar_width}" y="{y_pos + 14}" font-family="Inter" font-size="11px" font-weight="600" fill="#111827">R$ {val:,.2f}</text>
            """
        
        svg_chart = f"""
        <svg width="100%" height="{svg_height}" viewBox="0 0 {svg_width} {svg_height}" style="background-color: #F9FAFB; border-radius: 8px; border: 1px solid #E5E7EB; padding: 10px;">
            <text x="10" y="15" font-family="Inter" font-size="12px" font-weight="700" fill="#1E3A8A">VALUATION TOTAL POR SETOR (R$)</text>
            {chart_rows}
        </svg>
        """

    # 7. Renderizar Tabelas por Setor
    setores = sorted(df_estoque_ordenado["categoria"].unique())
    tabelas_setores_html = ""
    for setor in setores:
        df_setor = df_estoque_ordenado[df_estoque_ordenado["categoria"] == setor]
        linhas_html = ""
        setor_valuation = df_setor["valor_total"].sum()
        
        for _, row in df_setor.iterrows():
            nome = row["nome"]
            saldo = row["saldo_atual"]
            minimo = row["estoque_minimo"]
            preco = row["valor_unitario"]
            val_total = row["valor_total"]
            runway = row["Runway"]
            runway_txt = "Sem consumo" if runway == 999 else f"{runway} dias"
            abc_xyz = f"Classe {row['Classe_ABC']}-{row['criticidade']}"
            ponto_ped = row["Ponto_Pedido"]
            status = row["Status"]
            
            if status == "Ruptura":
                badge = '<span class="badge badge-danger">RUPTURA</span>'
            elif status == "Crítico":
                badge = '<span class="badge badge-warning" style="background-color:#FDE8E8; color:#9B1C1C;">CRÍTICO</span>'
            elif status == "Ponto de Pedido":
                badge = '<span class="badge badge-warning">PONTO DE PEDIDO</span>'
            else:
                badge = '<span class="badge badge-success">OK</span>'
                
            linhas_html += f"""
            <tr>
                <td><b>{nome}</b></td>
                <td style="text-align:center;">{saldo}</td>
                <td style="text-align:center;">{minimo}</td>
                <td style="text-align:center;">{ponto_ped}</td>
                <td style="text-align:center;">{runway_txt}</td>
                <td style="text-align:center;">R$ {preco:,.2f}</td>
                <td style="text-align:center;">R$ {val_total:,.2f}</td>
                <td style="text-align:center;"><b>{abc_xyz}</b></td>
                <td style="text-align:center;">{badge}</td>
            </tr>
            """
            
        tabelas_setores_html += f"""
        <div style="margin-top: 15px; background-color: #FFFFFF; border: 1px solid #E5E7EB; border-radius: 8px; overflow: hidden; margin-bottom: 25px;">
            <div style="background-color: #F9FAFB; padding: 10px 15px; border-bottom: 1px solid #E5E7EB; display: flex; justify-content: space-between; align-items: center;">
                <span style="font-weight: 700; color: #1E3A8A; font-size: 13px;">SETOR: {setor.upper()}</span>
                <span style="font-size: 12px; color: #4B5563;">Capital Imobilizado no Setor: <b>R$ {setor_valuation:,.2f}</b></span>
            </div>
            <table class="table" style="margin-bottom: 0;">
                <thead>
                    <tr>
                        <th>Insumo</th>
                        <th style="text-align:center;">Saldo</th>
                        <th style="text-align:center;">Mínimo</th>
                        <th style="text-align:center;">Pto. Ressuprimento</th>
                        <th style="text-align:center;">Cobertura</th>
                        <th style="text-align:center;">Preço Un.</th>
                        <th style="text-align:center;">Valor Total</th>
                        <th style="text-align:center;">Classe ABC-XYZ</th>
                        <th style="text-align:center;">Status</th>
                    </tr>
                </thead>
                <tbody>
                    {linhas_html}
                </tbody>
            </table>
        </div>
        """

    # 8. Renderizar Lista Automatizada de Compras
    df_compras = df_calc[df_calc["Sugestao_Compra"] > 0].copy()
    lista_compras_html = ""
    if not df_compras.empty:
        linhas_compras = ""
        total_compra = df_compras["Custo_Compra"].sum()
        for _, row in df_compras.iterrows():
            nome = row["nome"]
            saldo = row["saldo_atual"]
            min_ideal = row["Minimo_Ideal"]
            sug = row["Sugestao_Compra"]
            preco = row["valor_unitario"]
            custo = row["Custo_Compra"]
            classe = f"Classe {row['Classe_ABC']}-{row['criticidade']}"
            
            linhas_compras += f"""
            <tr>
                <td><b>{nome}</b></td>
                <td style="text-align:center;">{saldo} un</td>
                <td style="text-align:center;">{min_ideal} un</td>
                <td style="text-align:center; color:#E02424; font-weight:700;">{int(sug)} un</td>
                <td style="text-align:center;">R$ {preco:,.2f}</td>
                <td style="text-align:center; font-weight:600;">R$ {custo:,.2f}</td>
                <td style="text-align:center;">{classe}</td>
            </tr>
            """
            
        lista_compras_html = f"""
        <div class="section-title">2. Lista Automatizada de Compras (Reposição Recomendada WMS)</div>
        <div style="background-color: #FFFFFF; border: 1px solid #E5E7EB; border-radius: 8px; overflow: hidden; margin-bottom: 25px;">
            <table class="table" style="margin-bottom: 0;">
                <thead>
                    <tr>
                        <th>Insumo</th>
                        <th style="text-align:center;">Saldo Atual</th>
                        <th style="text-align:center;">Estoque Alvo</th>
                        <th style="text-align:center; color:#E02424;">Qtd. Recomendada</th>
                        <th style="text-align:center;">Preço Unit.</th>
                        <th style="text-align:center;">Total Estimado</th>
                        <th style="text-align:center;">Classe ABC-XYZ</th>
                    </tr>
                </thead>
                <tbody>
                    {linhas_compras}
                    <tr style="background-color: #F9FAFB; font-weight: bold; border-top: 2px solid #E5E7EB;">
                        <td colspan="4" style="text-align: right; padding-right: 20px;">INVESTIMENTO TOTAL PREVISTO EM COMPRAS:</td>
                        <td colspan="2" style="text-align: center; color: #1E3A8A; font-size: 14px;">R$ {total_compra:,.2f}</td>
                        <td></td>
                    </tr>
                </tbody>
            </table>
        </div>
        """
    else:
        lista_compras_html = """
        <div class="section-title">2. Lista Automatizada de Compras</div>
        <div style="background-color: #DEF7EC; color: #03543F; padding: 15px; border-radius: 8px; border: 1px solid #DEF7EC; font-size: 13px; font-weight: 500; margin-bottom: 25px;">
            🟢 **Nenhuma compra recomendada no momento.** Todos os itens encontram-se abastecidos e acima dos níveis mínimos de segurança.
        </div>
        """

    # 9. HTML do Resumo Executivo para Ação Imediata
    urgente_li = "".join([f"<li>{item}</li>" for item in urgente_comprar]) if urgente_comprar else "<li>Nenhum item em estado crítico ou ruptura.</li>"
    monitorar_li = "".join([f"<li>{item}</li>" for item in monitorar_aprovisionar]) if monitorar_aprovisionar else "<li>Nenhum item com risco iminente de ruptura.</li>"
    investigar_li = "".join([f"<li>{item}</li>" for item in investigar]) if investigar else "<li>Nenhum comportamento suspeito detectado (valuation e volumes normais).</li>"
    
    resumo_executivo_html = f"""
    <div style="background-color: #FFFFFF; border: 1px solid #E5E7EB; border-radius: 8px; padding: 20px; margin-bottom: 30px; box-shadow: 0 2px 4px rgba(0,0,0,0.02);">
        <div style="font-weight: 700; color: #1E3A8A; font-size: 14px; margin-bottom: 15px; border-bottom: 2px solid #E5E7EB; padding-bottom: 8px; text-transform: uppercase; letter-spacing: 0.5px;">
            🎯 Resumo Executivo & Plano de Ação Imediata
        </div>
        <div style="display: flex; gap: 15px;">
            <div style="flex: 1; background-color: rgba(239, 68, 68, 0.03); border-left: 4px solid #EF4444; padding: 12px; border-radius: 4px;">
                <span style="font-weight: 700; color: #E02424; font-size: 12px; text-transform: uppercase;">🚨 Comprar Urgentemente:</span>
                <ul style="font-size: 11px; padding-left: 15px; margin-top: 8px; color: #374151; line-height: 1.5; margin-bottom: 0;">
                    {urgente_li}
                </ul>
            </div>
            <div style="flex: 1; background-color: rgba(245, 158, 11, 0.03); border-left: 4px solid #F59E0B; padding: 12px; border-radius: 4px;">
                <span style="font-weight: 700; color: #D97706; font-size: 12px; text-transform: uppercase;">⚠️ Monitorar / Aprovisionar:</span>
                <ul style="font-size: 11px; padding-left: 15px; margin-top: 8px; color: #374151; line-height: 1.5; margin-bottom: 0;">
                    {monitorar_li}
                </ul>
            </div>
            <div style="flex: 1; background-color: rgba(59, 130, 246, 0.03); border-left: 4px solid #3B82F6; padding: 12px; border-radius: 4px;">
                <span style="font-weight: 700; color: #1D4ED8; font-size: 12px; text-transform: uppercase;">🔍 Investigar Alertas:</span>
                <ul style="font-size: 11px; padding-left: 15px; margin-top: 8px; color: #374151; line-height: 1.5; margin-bottom: 0;">
                    {investigar_li}
                </ul>
            </div>
        </div>
    </div>
    """

    # 10. Movimentações Recentes
    df_mov = mv.head(10).copy()
    df_mov = df_mov.rename(columns={
        "data_hora": "Data/Hora",
        "produto": "Insumo",
        "tipo": "Operação",
        "quantidade": "Qtd",
        "saldo_resultante": "Saldo Pos-Mov",
        "observacao": "Observação"
    })
    tb_mov_html = df_mov[["Data/Hora", "Insumo", "Operação", "Qtd", "Saldo Pos-Mov", "Observação"]].to_html(index=False, classes="table")

    data_geracao = datetime.datetime.now().strftime("%d/%m/%Y às %H:%M:%S")
    total_itens = len(df)
    
    # Construção do HTML Final
    html_content = f"""
    <!DOCTYPE html>
    <html lang="pt-BR">
    <head>
        <meta charset="UTF-8">
        <title>Relatório Executivo Gerencial WMS 5.0</title>
        <style>
            @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
            
            body {{
                font-family: 'Inter', sans-serif;
                color: #1F2937;
                background-color: #F9FAFB;
                margin: 0;
                padding: 30px;
                line-height: 1.5;
            }}
            
            .report-container {{
                max-width: 950px;
                margin: 0 auto;
                background-color: #FFFFFF;
                padding: 35px;
                border-radius: 12px;
                box-shadow: 0 4px 6px rgba(0,0,0,0.05);
                border: 1px solid #E5E7EB;
            }}
            
            .header-table {{
                width: 100%;
                border-bottom: 2px solid #1E3A8A;
                padding-bottom: 15px;
                margin-bottom: 25px;
            }}
            
            .logo-title {{
                color: #1E3A8A;
                font-size: 26px;
                font-weight: 700;
                margin: 0;
            }}
            
            .report-subtitle {{
                font-size: 13px;
                color: #6B7280;
                margin: 5px 0 0 0;
                text-transform: uppercase;
                letter-spacing: 1px;
            }}
            
            .meta-info {{
                text-align: right;
                font-size: 11px;
                color: #6B7280;
                line-height: 1.4;
            }}
            
            .section-title {{
                font-size: 16px;
                color: #1E3A8A;
                border-left: 4px solid #1E3A8A;
                padding-left: 10px;
                margin-top: 35px;
                margin-bottom: 15px;
                font-weight: 700;
                text-transform: uppercase;
                letter-spacing: 0.5px;
            }}
            
            /* Grid de KPIs */
            .kpi-container {{
                display: grid;
                grid-template-columns: repeat(4, 1fr);
                gap: 15px;
                margin-bottom: 25px;
            }}
            
            .kpi-card {{
                background-color: #F9FAFB;
                padding: 12px;
                border-radius: 8px;
                text-align: center;
                border: 1px solid #E5E7EB;
            }}
            
            .kpi-val {{
                font-size: 18px;
                font-weight: 700;
                color: #111827;
                margin: 4px 0 0 0;
            }}
            
            .kpi-val span {{
                font-size: 12px;
                font-weight: 500;
                color: #6B7280;
            }}
            
            .kpi-lbl {{
                font-size: 10px;
                color: #6B7280;
                text-transform: uppercase;
                font-weight: 600;
                letter-spacing: 0.5px;
            }}
            
            /* Tabelas */
            .table {{
                width: 100%;
                border-collapse: collapse;
                margin-bottom: 20px;
                font-size: 12px;
                text-align: left;
            }}
            
            .table th {{
                background-color: #F9FAFB;
                color: #374151;
                font-weight: 600;
                padding: 8px 10px;
                border-bottom: 2px solid #E5E7EB;
            }}
            
            .table td {{
                padding: 8px 10px;
                border-bottom: 1px solid #F3F4F6;
                color: #4B5563;
            }}
            
            /* Status Badges */
            .badge {{
                font-size: 9px;
                font-weight: 700;
                padding: 2px 6px;
                border-radius: 4px;
                text-transform: uppercase;
                letter-spacing: 0.5px;
            }}
            
            .badge-success {{
                background-color: #DEF7EC;
                color: #03543F;
            }}
            
            .badge-warning {{
                background-color: #FEF08A;
                color: #713F12;
            }}
            
            .badge-danger {{
                background-color: #FDE8E8;
                color: #9B1C1C;
            }}
            
            .footer-signature {{
                margin-top: 45px;
                width: 100%;
                font-size: 12px;
            }}
            
            .sig-line {{
                border-top: 1px solid #9CA3AF;
                width: 230px;
                margin-top: 35px;
                padding-top: 5px;
                text-align: center;
                color: #4B5563;
            }}
            
            @media print {{
                body {{
                    background-color: #FFFFFF;
                    padding: 0;
                    font-size: 11px;
                }}
                .report-container {{
                    border: none;
                    box-shadow: none;
                    padding: 0;
                    max-width: 100%;
                }}
                .no-print {{
                    display: none;
                }}
                .page-break {{
                    page-break-before: always;
                }}
            }}
        </style>
    </head>
    <body>
        <div class="report-container">
            <!-- Cabeçalho -->
            <table class="header-table" style="border-collapse: collapse;">
                <tr>
                    <td>
                        <h1 class="logo-title">📦 WMS 5.0</h1>
                        <p class="report-subtitle">Relatório Executivo WMS 5.0 - Gerencial & Auditoria</p>
                    </td>
                    <td class="meta-info">
                        <strong>Data de Emissão:</strong> {data_geracao}<br>
                        <strong>Abrangência:</strong> Almoxarifado Interno de Insumos
                    </td>
                </tr>
            </table>
            
            <!-- KPIs -->
            <div class="kpi-container">
                <div class="kpi-card">
                    <div class="kpi-lbl">Insumos Cadastrados</div>
                    <div class="kpi-val">{total_itens} itens</div>
                </div>
                <div class="kpi-card">
                    <div class="kpi-lbl">Valuation do Estoque</div>
                    <div class="kpi-val">R$ {total_valuation:,.2f}</div>
                </div>
                <div class="kpi-card" style="border-left: 3px solid #10B981;">
                    <div class="kpi-lbl">Acuracidade Inventário (IRA)</div>
                    <div class="kpi-val" style="color: {'#10b981' if ira >= 90 else ('#f59e0b' if ira >= 75 else '#ef4444')};">{ira:.1f}%</div>
                </div>
                <div class="kpi-card" style="border-left: 3px solid #EF4444;">
                    <div class="kpi-lbl">Financeiro de Perdas</div>
                    <div class="kpi-val" style="color: #9B1C1C;">R$ {perdas_financeiras:,.2f}</div>
                </div>
            </div>
            
            <!-- Resumo Executivo e Ações Imediatas -->
            {resumo_executivo_html}
            
            <div style="display: flex; gap: 15px; margin-bottom: 25px;">
                <div style="flex: 1.5;">
                    <!-- Lista de Compras -->
                    {lista_compras_html}
                </div>
                <div style="flex: 1; margin-top: 35px;">
                    <!-- Gráfico Valuation por Setor -->
                    {svg_chart}
                </div>
            </div>
            
            <div class="page-break"></div>
            
            <!-- Seção 1: Inventário -->
            <div class="section-title">1. Posição Consolidada por Setor</div>
            {tabelas_setores_html}
            
            <div class="page-break"></div>
            
            <!-- Seção 3: Movimentações -->
            <div class="section-title" style="margin-top: 30px;">3. Extrato Recente de Movimentações (Últimas 10)</div>
            <div style="background-color: #FFFFFF; border: 1px solid #E5E7EB; border-radius: 8px; overflow: hidden; margin-bottom: 25px;">
                {tb_mov_html}
            </div>
            
            <!-- Assinatura -->
            <table class="footer-signature">
                <tr>
                    <td style="vertical-align: bottom;">
                        <div class="sig-line">
                            Responsável Técnico Almoxarifado
                        </div>
                    </td>
                    <td style="text-align: right; font-size: 10px; color: #9CA3AF; vertical-align: bottom;">
                        Gerado eletronicamente via WMS 5.0 - Auditoria Operacional
                    </td>
                </tr>
            </table>
        </div>
    </body>
    </html>
    """
    return html_content

def gerar_html_diagnostico_ia(conteudo_md):
    """
    Converte o Markdown do diagnóstico da IA em uma página HTML premium otimizada para salvamento em PDF / Impressão.
    Utiliza codificação Base64 para tráfego seguro de caracteres especiais e tags.
    """
    import base64
    b64_str = base64.b64encode(conteudo_md.encode('utf-8')).decode('utf-8')
    
    html_content = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <title>Relatório de Diagnóstico IA - WMS 5.0</title>
    <script src="https://cdn.jsdelivr.net/npm/marked/marked.min.js"></script>
    <style>
        body {{
            font-family: 'Source Sans Pro', 'Inter', sans-serif;
            background-color: #f3f4f6;
            color: #1f2937;
            padding: 40px;
            margin: 0;
            display: flex;
            justify-content: center;
        }}
        .container {{
            background-color: #ffffff;
            border-radius: 16px;
            box-shadow: 0 10px 15px -3px rgba(0,0,0,0.05), 0 4px 6px -2px rgba(0,0,0,0.05);
            padding: 50px;
            width: 800px;
            border-top: 8px solid #1e3a8a;
        }}
        h1, h2, h3, h4, h5, h6 {{
            color: #1e3a8a;
            font-weight: 700;
            margin-top: 25px;
        }}
        h1 {{ border-bottom: 2px solid #e5e7eb; padding-bottom: 10px; }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
        }}
        th, td {{
            border: 1px solid #e5e7eb;
            padding: 12px;
            text-align: left;
        }}
        th {{
            background-color: #f3f4f6;
            font-weight: bold;
        }}
        ul, ol {{
            line-height: 1.6;
        }}
        .header-print {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            border-bottom: 2px solid #e5e7eb;
            padding-bottom: 20px;
            margin-bottom: 30px;
        }}
        .header-print .info {{
            text-align: right;
            font-size: 0.85rem;
            color: #6b7280;
        }}
        @media print {{
            body {{
                background-color: #ffffff;
                padding: 0;
            }}
            .container {{
                box-shadow: none;
                padding: 0;
                width: 100%;
                border-top: none;
            }}
            .btn-print {{
                display: none !important;
            }}
        }}
        .btn-print {{
            display: inline-block;
            background-color: #1e3a8a;
            color: white;
            padding: 12px 24px;
            border: none;
            border-radius: 8px;
            cursor: pointer;
            font-weight: bold;
            margin-bottom: 25px;
            font-size: 0.9rem;
            box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1);
            transition: background-color 0.2s;
        }}
        .btn-print:hover {{
            background-color: #1d4ed8;
        }}
    </style>
</head>
<body>
    <div class="container">
        <button class="btn-print" onclick="window.print()">🖨️ Imprimir / Salvar em PDF</button>
        <div class="header-print">
            <div>
                <h2 style="margin: 0; color: #1e3a8a;">WMS 5.0</h2>
                <span style="font-size: 0.85rem; color: #6b7280;">Relatório Preditivo do Assistente de Inteligência Artificial</span>
            </div>
            <div class="info">
                <strong>Data de Emissão:</strong> <span id="data-emissao"></span><br>
                <strong>Emitido por:</strong> Gestor do Sistema
            </div>
        </div>
        <div id="content"></div>
    </div>
    <script>
        document.getElementById('data-emissao').innerText = new Date().toLocaleDateString('pt-BR') + ' ' + new Date().toLocaleTimeString('pt-BR', {{hour: '2-digit', minute:'2-digit'}});
        const base64Text = "{b64_str}";
        const mdText = decodeURIComponent(escape(atob(base64Text)));
        document.getElementById('content').innerHTML = marked.parse(mdText);
    </script>
</body>
</html>"""
    return html_content
